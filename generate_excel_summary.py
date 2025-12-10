#!/usr/bin/env python3
"""
Generate Excel summary from t-test results.

Usage:
    python generate_excel_summary.py <base_output_dir>
"""

import json
import sys
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, numbers


def main():
    if len(sys.argv) < 2:
        print("Error: base_output_dir argument is required")
        print("Usage: python generate_excel_summary.py <base_output_dir>")
        sys.exit(1)
    
    base_output_dir = Path(sys.argv[1])
    
    if not base_output_dir.exists():
        print(f"Error: Directory not found: {base_output_dir}")
        sys.exit(1)
    
    # Collect all t-test results
    results_data = []
    
    eval_tasks = ["basic", "backforth"]
    target_frames = [21, 25]
    
    for eval_task in eval_tasks:
        for target_frame in target_frames:
            # Determine directory names based on new naming scheme
            # Format: instruct_pix2pix_224[_progress][_next5]_eval_{eval_task}
            if target_frame == 21:
                no_progress_dir = base_output_dir / f"instruct_pix2pix_224_eval_{eval_task}"
                progress_dir = base_output_dir / f"instruct_pix2pix_224_progress_eval_{eval_task}"
            else:
                # frame25 uses _next5 suffix
                no_progress_dir = base_output_dir / f"instruct_pix2pix_224_next5_eval_{eval_task}"
                progress_dir = base_output_dir / f"instruct_pix2pix_224_progress_next5_eval_{eval_task}"
            
            t_test_file = no_progress_dir / "t_test_results.json"
            
            if t_test_file.exists():
                with open(t_test_file, 'r', encoding='utf-8') as f:
                    t_test_data = json.load(f)
                
                # Extract SSIM results
                ssim = t_test_data['ssim']
                results_data.append({
                    'Eval Dataset': eval_task,
                    'Target Frame': target_frame,
                    'Metric': 'SSIM',
                    'Mean (with progress)': ssim['mean_with_progress'],
                    'Mean (without progress)': ssim['mean_without_progress'],
                    'Mean Difference': ssim['mean_difference'],
                    't-statistic': ssim['t_statistic'],
                    'p-value': ssim['p_value'],
                    'Significance': ssim['significance'],
                    'N (with progress)': ssim['n_with_progress'],
                    'N (without progress)': ssim['n_without_progress']
                })
                
                # Extract PSNR results
                psnr = t_test_data['psnr']
                results_data.append({
                    'Eval Dataset': eval_task,
                    'Target Frame': target_frame,
                    'Metric': 'PSNR',
                    'Mean (with progress)': psnr['mean_with_progress'],
                    'Mean (without progress)': psnr['mean_without_progress'],
                    'Mean Difference': psnr['mean_difference'],
                    't-statistic': psnr['t_statistic'],
                    'p-value': psnr['p_value'],
                    'Significance': psnr['significance'],
                    'N (with progress)': psnr['n_with_progress'],
                    'N (without progress)': psnr['n_without_progress']
                })
            else:
                print(f"Warning: t_test_results.json not found in {no_progress_dir}")
    
    if not results_data:
        print("Error: No t-test results found")
        sys.exit(1)
    
    # Create DataFrame
    df = pd.DataFrame(results_data)
    
    # Save to Excel
    excel_path = base_output_dir / "t_test_summary.xlsx"
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='T-Test Results', index=False)
        
        # Get workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['T-Test Results']
        
        # Style header row
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Format numeric columns
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            # Mean columns (4, 5, 6) - 2 decimal places
            for col_idx in [4, 5, 6]:
                cell = row[col_idx]
                if isinstance(cell.value, (int, float)):
                    cell.number_format = numbers.FORMAT_NUMBER_00
            # Statistics columns (7, 8) - 4 decimal places
            for col_idx in [7, 8]:
                cell = row[col_idx]
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "0.0000"
    
    print(f"Excel summary saved to: {excel_path}")
    print(f"Total rows: {len(df)}")
    
    return 0


if __name__ == "__main__":
    sys.exit(main())

